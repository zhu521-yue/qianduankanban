from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.responses import ApiError
from upload.models import ParsedFile
from upload.normalization import clean_header


def _validate_headers(headers: list[str]) -> tuple[str, ...]:
    if not headers or not any(headers):
        raise ApiError(400, "FILE_HEADER_EMPTY", "上传文件没有有效表头。")
    blank = [index + 1 for index, header in enumerate(headers) if not header]
    if blank:
        raise ApiError(400, "FILE_HEADER_BLANK", f"第{blank[:10]}列的表头为空。")
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise ApiError(400, "FILE_HEADER_DUPLICATE", f"文件存在重复表头：{', '.join(duplicates)}")
    return tuple(headers)


def _read_csv(content: bytes) -> ParsedFile:
    decoded: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ApiError(400, "FILE_ENCODING_INVALID", "CSV必须使用UTF-8或GB18030编码。")
    reader = csv.reader(io.StringIO(decoded))
    raw_headers = next(reader, [])
    headers = _validate_headers([clean_header(item) for item in raw_headers])
    rows: list[dict[str, Any]] = []
    for values in reader:
        padded = values + [None] * max(0, len(headers) - len(values))
        if any(value not in (None, "") for value in padded):
            rows.append(dict(zip(headers, padded, strict=False)))
    return ParsedFile(headers, tuple(rows))


def _read_xlsx(content: bytes) -> ParsedFile:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ApiError(400, "XLSX_INVALID", "无法读取XLSX文件。") from exc
    try:
        rows = workbook.active.iter_rows(values_only=True)
        raw_headers = list(next(rows, ()))
        headers = _validate_headers([clean_header(item) for item in raw_headers])
        result = [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value not in (None, "") for value in values)
        ]
        return ParsedFile(headers, tuple(result))
    finally:
        workbook.close()


def read_file(file_name: str, content: bytes) -> ParsedFile:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return _read_csv(content)
    if suffix == ".xlsx":
        return _read_xlsx(content)
    raise ApiError(400, "FILE_TYPE_INVALID", "仅支持.csv和.xlsx文件。")
