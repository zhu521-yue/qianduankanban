from __future__ import annotations

import hashlib
import json
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg


PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = PROJECT_ROOT / "8.11" / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.settings import get_settings  # noqa: E402


DATA_DIR = PROJECT_ROOT / "data" / "阿里巴巴"
SOURCE_FILES = [
    DATA_DIR / "2025.2.1-2026.1.31号阿里数据.xlsx",
    DATA_DIR / "2026.2.1-7.27阿里明细.xlsx",
]
EXPECTED_SOURCE_ROWS = {
    "2025.2.1-2026.1.31号阿里数据.xlsx": 671_472,
    "2026.2.1-7.27阿里明细.xlsx": 306_656,
}
EXPECTED_SHEET = "Sheet1"
OVERLAP_DATE = date(2026, 1, 31)
TARGET_DATABASE = "weidian"
TARGET_SCHEMA = "alibaba"
TARGET_TABLE = "raw_data"


# 使用第一份文件的字段顺序作为raw_data统一顺序；第二份文件按表头名称映射。
SOURCE_HEADERS = [
    "内部订单号",
    "标记多标签",
    "售后单号",
    "订单类型",
    "线上订单号",
    "订单状态",
    "发货仓",
    "虚拟仓",
    "分销商",
    "店铺",
    "小旗",
    "买家ID",
    "买家账号",
    "卖家备注",
    "订单日期",
    "发货日期",
    "付款日期",
    "业务员",
    "收货人",
    "省",
    "市",
    "区县",
    "快递公司",
    "快递单号",
    "售后登记日期",
    "售后确认日期",
    "售后分类",
    "问题类型",
    "商品编码",
    "款式编码",
    "原始线上订单号",
    "产品分类",
    "品牌",
    "商品简称",
    "颜色规格",
    "供应商",
    "店铺商品编码",
    "基本售价",
    "成本价",
    "成本价来源",
    "销售数量",
    "实发数量",
    "实发金额",
    "销售金额",
    "销售成本",
    "实发成本",
    "销售毛利",
    "销售毛利率",
    "已付金额",
    "应付金额",
    "售价",
    "基本金额",
    "退货数量",
    "实退数量",
    "退货金额",
    "退货成本",
    "实退成本",
    "实退金额",
    "运费收入",
    "运费收入分摊",
    "整单运费支出",
    "运费支出",
    "优惠金额",
    "订单重量",
    "供销商",
    "币种",
    "国际单号",
    "剩余发货时间",
    "付款方式",
    "买家指定物流",
    "收货国家地区",
    "实发物流渠道",
    "境外运费支出分摊",
    "境外收入总计分摊",
    "境外支出总计分摊",
    "汇率",
    "货主分销",
    "平台补贴金额",
]


def quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def raw_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def parse_payment_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for pattern in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"无法解析付款日期：{value!r}")


def canonical_row_hash(values: list[Any]) -> str:
    payload = json.dumps(
        [raw_text(value) for value in values],
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def preflight_files() -> None:
    official_files = sorted(path for path in DATA_DIR.glob("*.xlsx") if not path.name.startswith("~$"))
    if official_files != sorted(SOURCE_FILES):
        raise RuntimeError(f"正式XLSX文件清单与设计不一致：{[path.name for path in official_files]}")
    for path in SOURCE_FILES:
        if not path.is_file():
            raise FileNotFoundError(path)


def workbook_rows(path: Path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != [EXPECTED_SHEET]:
            raise RuntimeError(f"{path.name}工作表不符合预期：{workbook.sheetnames}")
        worksheet = workbook[EXPECTED_SHEET]
        iterator = worksheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        if len(headers) != 78 or len(set(headers)) != 78 or set(headers) != set(SOURCE_HEADERS):
            missing = sorted(set(SOURCE_HEADERS) - set(headers))
            extra = sorted(str(value) for value in set(headers) - set(SOURCE_HEADERS))
            raise RuntimeError(f"{path.name}表头不符合预期，缺少={missing}，新增={extra}")
        header_index = {name: index for index, name in enumerate(headers)}
        for source_row_number, row in enumerate(iterator, start=2):
            yield source_row_number, [row[header_index[name]] for name in SOURCE_HEADERS]
    finally:
        workbook.close()


def create_table_sql() -> str:
    source_columns = ",\n".join(f"    {quote_identifier(header)} TEXT" for header in SOURCE_HEADERS)
    return f"""
CREATE TABLE {TARGET_SCHEMA}.{TARGET_TABLE} (
    id BIGSERIAL PRIMARY KEY,
{source_columns},
    created_at TIMESTAMP NOT NULL DEFAULT timezone('Asia/Shanghai', CURRENT_TIMESTAMP),
    updated_at TIMESTAMP NOT NULL DEFAULT timezone('Asia/Shanghai', CURRENT_TIMESTAMP)
)
"""


def main() -> None:
    if len(SOURCE_HEADERS) != 78:
        raise AssertionError(f"原始字段不是78列：{len(SOURCE_HEADERS)}")
    preflight_files()

    settings = get_settings()
    quoted_columns = ", ".join(quote_identifier(header) for header in SOURCE_HEADERS)
    copy_sql = f"COPY {TARGET_SCHEMA}.{TARGET_TABLE} ({quoted_columns}) FROM STDIN"
    source_rows: Counter[str] = Counter()
    inserted_rows: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()
    overlap_hashes: Counter[str] = Counter()
    distinct_buyers: set[str] = set()
    distinct_products: set[str] = set()
    blank_sales_quantity = 0
    blank_refund_quantity = 0
    blank_refund_amount = 0
    exact_duplicates_removed = 0
    payment_index = SOURCE_HEADERS.index("付款日期")
    status_index = SOURCE_HEADERS.index("订单状态")
    buyer_index = SOURCE_HEADERS.index("买家ID")
    product_index = SOURCE_HEADERS.index("商品编码")
    sales_quantity_index = SOURCE_HEADERS.index("销售数量")
    refund_quantity_index = SOURCE_HEADERS.index("实退数量")
    refund_amount_index = SOURCE_HEADERS.index("实退金额")

    print("阶段1/3：开启事务并将92列表替换为81列表", flush=True)
    with psycopg.connect(
        settings.database_url,
        connect_timeout=settings.database_connect_timeout,
    ) as connection:
        try:
            connection.execute("SET statement_timeout = 0")
            connection.execute("SELECT pg_advisory_xact_lock(hashtext(%s))", (f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}",))
            current = connection.execute(
                """
                SELECT
                    EXISTS (SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s),
                    (SELECT COUNT(*) FROM information_schema.columns WHERE table_schema=%s AND table_name=%s)
                """,
                (TARGET_SCHEMA, TARGET_TABLE, TARGET_SCHEMA, TARGET_TABLE),
            ).fetchone()
            if not current[0] or current[1] != 92:
                raise RuntimeError(f"当前目标不是预期的92列表，拒绝重建：exists={current[0]}，columns={current[1]}")
            current_rows = connection.execute(f"SELECT COUNT(*) FROM {TARGET_SCHEMA}.{TARGET_TABLE}").fetchone()[0]
            if current_rows != 978_128:
                raise RuntimeError(f"当前92列表行数不是978,128，拒绝重建：{current_rows:,}")

            connection.execute(f"DROP TABLE {TARGET_SCHEMA}.{TARGET_TABLE}")
            connection.execute(create_table_sql())

            print("阶段2/3：流式读取两份Excel并写入78个原始字段", flush=True)
            with connection.cursor() as cursor:
                with cursor.copy(copy_sql) as copy:
                    for file_index, path in enumerate(SOURCE_FILES):
                        print(f"开始读取：{path.name}", flush=True)
                        for source_row_number, raw_values in workbook_rows(path):
                            source_rows[path.name] += 1
                            try:
                                payment_date = parse_payment_date(raw_values[payment_index])
                            except Exception as exc:
                                raise RuntimeError(f"{path.name}第{source_row_number}行付款日期错误：{exc}") from exc

                            if payment_date == OVERLAP_DATE:
                                row_hash = canonical_row_hash(raw_values)
                                if file_index == 0:
                                    overlap_hashes[row_hash] += 1
                                elif overlap_hashes[row_hash] > 0:
                                    overlap_hashes[row_hash] -= 1
                                    exact_duplicates_removed += 1
                                    continue

                            converted = tuple(raw_text(value) for value in raw_values)
                            copy.write_row(converted)
                            inserted_rows[path.name] += 1

                            status = converted[status_index]
                            if status is not None:
                                status_counts[status] += 1
                            buyer_id = converted[buyer_index]
                            product_code = converted[product_index]
                            if buyer_id:
                                distinct_buyers.add(buyer_id)
                            if product_code:
                                distinct_products.add(product_code)
                            blank_sales_quantity += int(is_blank(converted[sales_quantity_index]))
                            blank_refund_quantity += int(is_blank(converted[refund_quantity_index]))
                            blank_refund_amount += int(is_blank(converted[refund_amount_index]))

                            if source_rows[path.name] % 50_000 == 0:
                                print(
                                    f"{path.name}已读取{source_rows[path.name]:,}行，已写入{inserted_rows[path.name]:,}行",
                                    flush=True,
                                )

                        expected_rows = EXPECTED_SOURCE_ROWS[path.name]
                        if source_rows[path.name] != expected_rows:
                            raise RuntimeError(
                                f"{path.name}源行数不符：实际{source_rows[path.name]:,}，预期{expected_rows:,}"
                            )
                        print(
                            f"完成读取：{path.name}，源数据{source_rows[path.name]:,}行，写入{inserted_rows[path.name]:,}行",
                            flush=True,
                        )

            print("阶段3/3：执行81列结构与原始数据对账", flush=True)
            expected_columns = ["id", *SOURCE_HEADERS, "created_at", "updated_at"]
            database_columns = [
                row[0]
                for row in connection.execute(
                    "SELECT column_name FROM information_schema.columns WHERE table_schema=%s AND table_name=%s ORDER BY ordinal_position",
                    (TARGET_SCHEMA, TARGET_TABLE),
                )
            ]
            if database_columns != expected_columns:
                raise RuntimeError("数据库81列的名称或顺序不符合预期")

            summary = connection.execute(
                f"""
                SELECT
                    COUNT(*),
                    COUNT(DISTINCT id),
                    MIN(id),
                    MAX(id),
                    COUNT(*) FILTER (WHERE {quote_identifier('订单状态')} = '已发货'),
                    COUNT(*) FILTER (WHERE {quote_identifier('订单状态')} = '已取消'),
                    COUNT(*) FILTER (WHERE {quote_identifier('销售数量')} IS NULL OR BTRIM({quote_identifier('销售数量')}) = ''),
                    COUNT(*) FILTER (WHERE {quote_identifier('实退数量')} IS NULL OR BTRIM({quote_identifier('实退数量')}) = ''),
                    COUNT(*) FILTER (WHERE {quote_identifier('实退金额')} IS NULL OR BTRIM({quote_identifier('实退金额')}) = ''),
                    COUNT(DISTINCT {quote_identifier('买家ID')}),
                    COUNT(DISTINCT {quote_identifier('商品编码')}),
                    COUNT(*) FILTER (WHERE created_at IS NULL OR updated_at IS NULL),
                    MIN(created_at),
                    MAX(created_at)
                FROM {TARGET_SCHEMA}.{TARGET_TABLE}
                """
            ).fetchone()

            expected_inserted = sum(inserted_rows.values())
            expected_statuses = {"已发货": status_counts["已发货"], "已取消": status_counts["已取消"]}
            actual_statuses = {"已发货": summary[4], "已取消": summary[5]}
            validations = [
                (summary[0] == expected_inserted, "总行数"),
                (summary[1] == expected_inserted and summary[2] == 1 and summary[3] == expected_inserted, "自增主键"),
                (actual_statuses == expected_statuses, "订单状态"),
                (summary[6] == blank_sales_quantity, "空销售数量"),
                (summary[7] == blank_refund_quantity, "空实退数量"),
                (summary[8] == blank_refund_amount, "空实退金额"),
                (summary[9] == len(distinct_buyers), "买家ID"),
                (summary[10] == len(distinct_products), "商品编码"),
                (summary[11] == 0, "北京时间技术字段"),
            ]
            failed = [name for passed, name in validations if not passed]
            if failed:
                raise RuntimeError(f"事务内对账失败：{failed}")

            connection.commit()
            result = {
                "database": TARGET_DATABASE,
                "schema": TARGET_SCHEMA,
                "table": TARGET_TABLE,
                "column_count": len(database_columns),
                "source_rows": dict(source_rows),
                "inserted_rows": dict(inserted_rows),
                "cross_file_exact_duplicates_removed": exact_duplicates_removed,
                "database_rows": summary[0],
                "shipped_rows": summary[4],
                "cancelled_rows": summary[5],
                "blank_sales_quantity": summary[6],
                "blank_refund_quantity": summary[7],
                "blank_refund_amount": summary[8],
                "distinct_buyers": summary[9],
                "distinct_products": summary[10],
                "created_at_min": str(summary[12]),
                "created_at_max": str(summary[13]),
            }
            print("REBUILD_RESULT=" + json.dumps(result, ensure_ascii=False), flush=True)
        except Exception:
            connection.rollback()
            print("重建失败，事务已整体回滚，原92列表已恢复。", flush=True)
            raise


if __name__ == "__main__":
    main()
